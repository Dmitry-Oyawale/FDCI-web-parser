from app import db
from flask import render_template, flash, redirect, url_for, request, jsonify, send_file, current_app
import uuid
import sqlalchemy as sqla

from app.src import src_blueprint as src
from app.src.forms import ParseForm

import requests
from bs4 import BeautifulSoup #pip install bs4
import pandas
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
from io import StringIO
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
from openpyxl.styles import Font

BIG = Font(size=14, bold=True)
SMALL = Font(size=11)

def get_rendered_html(url: str) -> str:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(user_agent="Mozilla/5.0")

        page.route("**/*", lambda route: route.abort()
                   if route.request.resource_type in ("image", "media", "font")
                   else route.continue_())

        page.goto(url, wait_until="domcontentloaded", timeout=60000)

        selector = "div#jumptocontent.teacher-page-content"
        page.wait_for_selector(selector, timeout=60000)

        page.wait_for_function(
            """(sel) => {
                const el = document.querySelector(sel);
                return el && el.innerText && el.innerText.trim().length > 50;
            }""",
            arg=selector,
            timeout=60000
        )

        html = page.content()
        browser.close()
        return html


@src.route('/', methods=['GET', 'POST'])
@src.route('/index', methods=['GET', 'POST'])
def index():
    return redirect(url_for('src.parse'))

@src.route('/parse', methods=['GET', 'POST'])
def parse():   
    form = ParseForm()
    if form.validate_on_submit():
        if form.link.data and form.file.data:

            url = form.link.data
            rendered_html = get_rendered_html(url)
            soup = BeautifulSoup(rendered_html, "html.parser")

            target_div = soup.select_one("div#jumptocontent.teacher-page-content")
            if not target_div or not target_div.get_text(strip=True):
                target_div = soup.select_one("div.teacher-page-content")

            if not target_div:
                flash("Could not find lesson content after render.", "warning")
                return redirect(url_for("src.parse"))
            
            blocks = []
            for el in target_div.select("h1,h2,h3,h4,h5,h6,p,li"):
                text = el.get_text(" ", strip=True)
                if not text:
                    continue

                is_big = el.name in {"h1", "h2", "h3", "h4", "h5", "h6"}
                if el.name == "li":
                    text = f"• {text}"

                blocks.append({"Text": text, "IsBig": is_big})

            if not blocks:
                flash("No readable text blocks found.", "warning")
                return redirect(url_for("src.parse"))

            data_frame = pandas.DataFrame(blocks)
            print(data_frame.head(5))

            upload = form.file.data
            filename = secure_filename(upload.filename)

            if not filename.lower().endswith(".xlsx"):
                flash("Please upload an .xlsx file.", "warning")
                return redirect(url_for("src.parse"))

            upload_dir = os.path.join(current_app.instance_path, "uploads")
            os.makedirs(upload_dir, exist_ok=True)
            unique_name = f"{uuid.uuid4().hex}_{filename}"

            save_path = os.path.join(upload_dir, unique_name)
            upload.save(save_path)

            sheet_name = "Sheet1"

            wb = load_workbook(save_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            is_empty = (ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None)
            start_row = 1 if is_empty else ws.max_row + 1
            
            with pandas.ExcelWriter(save_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_frame.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=(start_row == 1),
                    startrow=start_row - 1
                )

            wb = load_workbook(save_path)
            ws = wb[sheet_name]

            header_rows = 1 if (start_row == 1) else 0
            first_data_row = start_row + header_rows
            last_data_row = first_data_row + len(data_frame) - 1

            for r in range(first_data_row, last_data_row + 1):
                is_big = bool(ws.cell(row=r, column=2).value)
                ws.cell(row=r, column=1).font = BIG if is_big else SMALL

            ws.delete_cols(2)  

            wb.save(save_path)

            return send_file(save_path, as_attachment=True, download_name=filename)


        else: 
            flash('Please provide both a link and file.', 'warning')
            return redirect(url_for('src.parse'))
    return render_template('parse.html', form=form)